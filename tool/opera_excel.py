# 操作excel
from openpyxl import load_workbook
from openpyxl import workbook


class OperaExcel():
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    def getRowsNum(self):
        """获取表格总行数"""
        rows = self.ws.max_row
        return rows

    def getColsNum(self):
        """获取表格总列数"""
        cols = self.ws.max_column
        return cols

    def getCellValue(self, row, column):
        """获取某个单元格的值"""
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    def getColValues(self, column):
        """获取某列的所有值——列表"""
        rows = self.ws.max_row
        columndata = []
        for row in range(2, rows + 1):
            cellvalue = self.ws.cell(row=row, column=column).value
            columndata.append(cellvalue)
        return columndata

    def getRowValues(self, row):
        """获取某行的所有值——列表"""
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata
